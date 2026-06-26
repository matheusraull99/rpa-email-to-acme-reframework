"""Gera os Config.xlsx (padrao REFramework) do Dispatcher e do Performer."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    style_header(ws, len(headers))
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 14), 60)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


# ---------- abas comuns ----------
SETTINGS_HEADERS = ["Name", "Value", "Description"]

COMMON_SETTINGS = [
    ["OrchestratorQueueName", "EmailWorkItems", "Nome da Queue no Orchestrator"],
    ["logF_BusinessProcessName", "EmailToACME", "Nome do processo nos logs"],
]

CONSTANTS_HEADERS = ["Name", "Value", "Description"]
COMMON_CONSTANTS = [
    ["MaxRetryNumber", 2, "Tentativas por transacao em System Exception"],
    ["MaxConsecutiveSystemExceptions", 3, "Aborta apos N falhas seguidas de sistema"],
    ["RetryNumberGetTransactionItem", 2, "Retries ao buscar item da fila"],
    ["RetryNumberSetTransactionStatus", 2, "Retries ao gravar status na fila"],
    ["ShouldMarkJobAsFaulted", "False", "Marca job como Faulted se terminar com erro"],
]

ASSETS_HEADERS = ["Name", "AssetName", "Description"]


def build(path, settings_extra, assets):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settings"
    write_sheet(ws, SETTINGS_HEADERS, COMMON_SETTINGS + settings_extra)

    ws2 = wb.create_sheet("Constants")
    write_sheet(ws2, CONSTANTS_HEADERS, COMMON_CONSTANTS)

    ws3 = wb.create_sheet("Assets")
    write_sheet(ws3, ASSETS_HEADERS, assets)

    wb.save(path)
    print("OK ->", path)


# ---------- Dispatcher ----------
build(
    "../1.Dispatcher/Data/Config.xlsx",
    settings_extra=[
        ["EmailAccount", "rpa.portfolio.teste@gmail.com", "Conta de e-mail monitorada (teste)"],
        ["EmailReadFolder", "Inbox", "Pasta de leitura"],
        ["EmailProcessedFolder", "Processed", "Pasta destino dos e-mails tratados"],
        ["EmailRejectedFolder", "Processed/Rejected", "Pasta para e-mails invalidos"],
        ["MaxEmailsPerRun", "50", "Limite de e-mails por execucao"],
        ["AcceptedVendors", "ACME Corp;Lisper;Saturn;Gold Trust", "Fornecedores aceitos (separados por ;)"],
    ],
    assets=[],
)

# ---------- Performer ----------
build(
    "../2.Performer/Data/Config.xlsx",
    settings_extra=[
        ["ACME_URL", "https://acme-test.uipath.com", "URL do sistema ACME System 1"],
        ["ACME_Credential_Asset", "ACME_Credential", "Asset (Credential) do Orchestrator com login ACME"],
    ],
    assets=[
        ["ACME_Credential", "ACME_Credential", "Credential do Orchestrator (usuario/senha ACME)"],
    ],
)
